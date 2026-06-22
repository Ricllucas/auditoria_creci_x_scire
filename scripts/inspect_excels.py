import json
from pathlib import Path

from openpyxl import load_workbook


FILES = [
    Path(r"C:\Users\ricll\Downloads\Base_Scire_melhorias .xlsx"),
    Path(r"C:\Users\ricll\Downloads\chamados_sead.xlsx"),
    Path(r"C:\Users\ricll\Downloads\chamados_suportes_fiscalizacao.xlsx"),
    Path(r"C:\Users\ricll\Downloads\chamados_suportes_processos_disciplinares.xlsx"),
    Path(r"C:\Users\ricll\Downloads\chamados_suportes_procuradoria_fiscal.xlsx"),
    Path(r"C:\Users\ricll\Downloads\chamados_suportes_sead.xlsx"),
]


def stringify(value):
    if value is None:
        return ""
    return str(value).strip()


def preview_rows(ws, limit=8, max_cols=12):
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(limit, ws.max_row), values_only=True):
        values = [stringify(value) for value in row[:max_cols]]
        if any(values):
            rows.append(values)
    return rows


def header_candidates(ws, max_header_row=6, preview_rows_count=2, max_cols=15):
    candidates = []
    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_header_row + preview_rows_count + 1), values_only=True))
    for header_index in range(min(max_header_row, len(rows))):
        header = [stringify(value) or f"__empty_{idx}" for idx, value in enumerate(rows[header_index][:max_cols])]
        data_preview = []
        for row in rows[header_index + 1 : header_index + 1 + preview_rows_count]:
            record = {}
            for idx, column in enumerate(header):
                value = stringify(row[idx] if idx < len(row) else "")
                if value:
                    record[column] = value
            if record:
                data_preview.append(record)
        candidates.append(
            {
                "header_row": header_index,
                "columns": header,
                "rows_preview": data_preview,
            }
        )
    return candidates


def summarize_workbook(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return {
            "file": path.name,
            "sheets": [
                {
                    "sheet": ws.title,
                    "shape": [ws.max_row, ws.max_column],
                    "top_rows": preview_rows(ws),
                    "header_candidates": header_candidates(ws),
                }
                for ws in workbook.worksheets
            ],
        }
    finally:
        workbook.close()


def main():
    summaries = []
    for path in FILES:
        if not path.exists():
            continue
        try:
            summaries.append(summarize_workbook(path))
        except Exception as exc:
            summaries.append({"file": path.name, "error": str(exc)})
    print(json.dumps(summaries, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()