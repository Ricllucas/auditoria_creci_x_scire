import openpyxl, re

filepath = r"C:\Users\ricll\Downloads\Base_Scire_melhorias .xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    all_rows = list(ws.iter_rows(values_only=True))
    print(f"SHEET: '{sheetname}' total rows={len(all_rows)}")

    print("\n--- TODOS OS VALORES DE COL6 (linha, col0, col6, type) ---")
    for i, row in enumerate(all_rows):
        col0 = str(row[0]).strip() if row[0] is not None else ''
        col6 = row[6] if len(row) > 6 else 'N/A'
        print(f"  row[{i:3d}] col0={repr(col0[:30]):<35} col6={repr(col6):<20} type={type(col6).__name__}")

wb.close()
